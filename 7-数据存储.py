'''
import json

s = '{"name":"张三"}'

# 将字符串转成字典对象
obj = json.loads(s) # 把JSON格式字符串转换成python字典
print(type(obj))
print(obj)

# 将字典对象转成字符串
obj = json.dumps(obj,ensure_ascii=False) # 把python字典转换成JSON格式字符串
print(type(obj))
print(obj) # 不加ensure_ascii=False则显示中文的unicode编码

# 把字典对象保存在文件中
json.dump(obj,open('json-try.txt','w',encoding='utf-8'),ensure_ascii=False)

# 读取文件中的json字符串
obj = json.load(open('json-try.txt',encoding='utf-8'))
print(obj)

# json爬取京东粽子销售情况
import json,requests
# 评价页面F12->Network,找到有Comment字样的项的URL
url = 'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=7252347&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&fold=1'
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36 Edg/101.0.1210.47'}
resp = requests.get(url,headers=headers)
text = resp.text
s = text.replace('fetchJSON_comment98(','').replace(');','') # 观察可知返回页面需要除去开始和结尾的几个字符才是完整的json字符串
#print(s)
s = json.loads(s) # 把字符串转化成字典对象
json.dump(s,open('JD-Zongzi-Sell-Best.txt','w',encoding='utf-8'),ensure_ascii=False)
comments = s['comments']
for comment in comments:
    content = comment['content'] # 评论内容文本
    time = comment['creationTime'] # 评论时间
    print(content, time)

# CSV读
import csv
with open('student.csv','r', encoding='utf-8',newline='\n') as f:
    obj = csv.reader(f) # 创建写入对象
    for row in obj:
        print(row)
# CSV写
with open('student.csv','a+', encoding='utf-8',newline='\n') as f:
    obj = csv.writer(f) # 创建写入对象
    obj.writerow(['赵六',21,88]) # 写入一行文件
    lst = [['王七',22,95],
    ['吴八',21,56],
    ['孙九',21,78]]
    obj.writerows(lst) # 一次写入多行数据

#使用CSV文件存储京东粽子评论数据
import json,requests,csv
# 评价页面F12->Network,找到有Comment字样的项的URL
url = 'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=7252347&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&fold=1'
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36 Edg/101.0.1210.47'}
resp = requests.get(url,headers=headers)
text = resp.text
s = text.replace('fetchJSON_comment98(','').replace(');','') # 观察可知返回页面需要除去开始和结尾的几个字符才是完整的json字符串
# print(s)
s = json.loads(s) # 把字符串转化成字典对象
# json.dump(s,open('JD-Zongzi-Sell-Best.txt','w',encoding='utf-8'),ensure_ascii=False)
# 观察json文件可知评论内容文本和评论时间各自的键值对
comments = s['comments']
save_list = []
for comment in comments:
    content = comment['content'] # 评论内容文本
    time = comment['creationTime'] # 评论时间
    print(content, time)
    save_list.append([content, time]) # 存储评论数据
# 写入CSV文件
with open('JD-Zongzi-Sell-Best.csv','w',newline='\n') as f:
    obj = csv.writer(f)
    obj.writerows(save_list)

# openpyxl基本操作
# 写入
import openpyxl 
obj = openpyxl.Workbook() # 创建对象
sheet = obj.active # 获取工作表sheet
cell = sheet['A1'] # 获取1行A列的指定单元格cell
cell.value = 'China' # 写入一个单元格
lst = ['Germany','Russia','Japan']
sheet.append(lst) # 写入一行数据 (写入多行可以通过循环实现)
obj.save('openpyxl-1.xlsx')

# 读取
import openpyxl
obj = openpyxl.load_workbook('openpyxl-1.xlsx') # 创建对象
sheet = obj.active # 仅限于文件中只有一个sheet(一个Excel文件可能有多个工作表)
sheet = obj['Sheet'] # 可用于文件中含有多个sheet的情况
cell = sheet['A2'] # 获取指定单元格
value = cell.value # 获取指定单元格的值
print(value)
columns = sheet['A'] # 获取一系列单元格,还可以sheet[3]/sheet['A:C'] 
for col in columns:
    print(col.value) # 获取一系列单元格的值

# 爬取下厨房的菜品数据
import requests,openpyxl
from bs4 import BeautifulSoup
# 评价页面F12->Network,找到有Comment字样的项的URL
url = 'https://www.xiachufang.com/explore/'
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36 Edg/101.0.1210.47'}
resp = requests.get(url,headers=headers)
text = resp.text
# print(text)
bs = BeautifulSoup(text,'lxml')
names = bs.find_all('p',class_='name') # 菜品名字
category = bs.find_all('p',class_='ing ellipsis') # 菜单
result = []
for i in range(len(names)):
    link = 'https://www.xiachufang.com/e'+names[i].find('a')['href']
    result.append([i+1,names[i].text[18:-14],category[i].text[1:-1],link],)
print(result)
# 保存
obj = openpyxl.Workbook()
sheet = obj.active
for row in result:
    sheet.append(row)
obj.save('下厨房菜品数据.xlsx')

# 操作MySQL
import mysql.connector
# 创建连接对象
conn = mysql.conector.connect (host='localhost', user='root', passwd='root', database='test , auth_plugin='mysql_native_password')#print (conn)
print(conn)
# 获取游标
mycursor = conn.cursor() 
# 编写sql语句
sql='insert into dept (deptno,dname,loc) values (%s,%s,%s)'
val=(50,'开发部','北京')
# 执行sql语句
mycursor.execute(sql,val)
# 提交
conn.commit()
print(mycursor.rowcount,'记录插入成功!')
'''